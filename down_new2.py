from pytube import YouTube
import pandas as pd
from openpyxl import load_workbook
import scrapetube
from pytube import YouTube
from datetime import datetime
import os
from dateutil.parser import parse
import sys
from ftplib import FTP
import ftplib
import shutil
from pytube import innertube
import time
import openpyxl
start_time = None
debug = True
#여기도 있다면?
## 위에 내용추가합니다~

def reset_timer(stream, chunk, bytes_remaining):
    global start_time
    if start_time is None:
        start_time = time.time()
    else:
        time_elapsed = time.time() - start_time
        if time_elapsed > 15:
            raise Exception('다운로드가 너무 오래 걸림')

def dt(text=None,log=True,is_return=False):
    t = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if is_return==True:
        return t
    if log==True:
        if text==None:
            print(t)
        else:
            print(t, text)
            
def ytdownload(file):
    innertube._cache_dir = os.path.join(os.getcwd(), "cache")
    innertube._token_file = os.path.join(innertube._cache_dir, 'tokens.json')

    error_yes = False
    wb = load_workbook(file)
    sheet_or= wb.worksheets[0]
    try:
        workbook = load_workbook('youtube_list.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        new_data = [
            '채널명',
            '영상링크',
            '파일명',
            '다운로드완료체크'
            '다운로드실패사유',
            '화질',
            '형식',
            '영상생성날짜',
            '영상 길이'
            ]
        sheet.append(new_data)
    workbook.save('youtube_list.xlsx')
    workbook.close()

    dt("영상 다운로드 프로세스 시작",debug)
    # print('st')
    for i, row in enumerate(sheet_or.rows):
        # print('sss')
        if i ==0:
            continue
        vurl = row[2].value
        videoid = vurl.split('/')[-1]
        first_filename = videoid + '.mp4'
        # print(videoid)

        ##

        if vurl == None:
            break
        print(i,vurl)
        start_time = datetime.now()
        download_val = ''
        resol = ''
        first_format = ''
        error_message = '정상 다운로드'
        channel_name = row[1].value

        try:
            yt = YouTube(vurl.strip(), use_oauth=True, allow_oauth_cache=True)
            if i == 1:
                pre_channel = yt.author
            ## 영상 저장 폴더만들기
            if not os.path.exists('youtube_video'):
                os.makedirs('youtube_video')
            fname = './youtube_video/'+videoid+'.mp4'
            message = str(yt.title) + ' 영상 다운로드 시작'
            dt(message,debug)
            # print(fname)

            download_session = yt.streams.filter(progressive=True, file_extension='mp4', res='720p' or '1080p')

            if len(download_session)==0:
                print('다운로드 대상 없음')

                error_message = '고화질 영상이 없음'
                workbook = openpyxl.load_workbook('youtube_list.xlsx')
                sheet = workbook.active
                new_data = [
                    yt.author,
                    vurl,
                    first_filename,
                    download_val,
                    error_message,
                    resol,
                    first_format,
                    yt.publish_date.strftime("%Y-%m-%d"),
                    yt.length
                    ]
                sheet.append(new_data)
                workbook.save('youtube_list.xlsx')
                workbook.close()

            else:
                start_time = time.time()
                timeout_seconds = 15
                timeout_chk = 0
                download_success = False
                while timeout_chk <3:
                    try:
                        resol =download_session.resolution
                        dfs = download_session[0].download(filename=fname)
                        time_elapsed = time.time() - start_time
                        if time_elapsed > timeout_seconds:
                            raise Exception('다운로드 시간 초과')
                        workbook = openpyxl.load_workbook('youtube_list.xlsx')
                        sheet = workbook.active
                        error_message = '없음'
                        download_val = 'O'
                        first_format = 'MP4'
                        new_data = [
                            yt.author,
                            vurl,
                            first_filename,
                            download_val,
                            error_message,
                            resol,
                            first_format,
                            yt.publish_date.strftime("%Y-%m-%d"),
                            yt.length
                            ]
                        sheet.append(new_data)
                        workbook.save('youtube_list.xlsx')
                        workbook.close()
                        download_success = True
                        break
                    except Exception as e:
                        print(f'다운로드에러 : {e}')
                        # message2 = e
                        # dt(message2,debug)
                        timeout_chk += 1
                        # error_message = '영상다운로드시간초과'

                if download_success == False:
                    error_message = '다운로드 시간 초과'
                    workbook = openpyxl.load_workbook('youtube_list.xlsx')
                    sheet = workbook.active
                    new_data = [
                        yt.author,
                        vurl,
                        first_filename,
                        download_val,
                        error_message,
                        resol,
                        first_format,
                        yt.publish_date.strftime("%Y-%m-%d"),
                        yt.length
                        ]
                    sheet.append(new_data)
                    workbook.save('youtube_list.xlsx')
                    workbook.close()
                    download_success = True
                    print("다운로드가 실패한 듯 보입니다. 다음영상으로 넘어갑니다.")
                    start_time = None
        except:
            error_message = '영상 확보 실패'
            workbook = openpyxl.load_workbook('youtube_list.xlsx')
            sheet = workbook.active
            new_data = [
                channel_name,
                vurl,
                first_filename,
                download_val,
                error_message,
                resol,
                first_format,
                '',
                ''
                ]
            sheet.append(new_data)
            workbook.save('youtube_list.xlsx')
            workbook.close()
            continue
# 둘다별로임
## 바뀐거 알고싶다
## 아 diff어떻게  쓰는거야

file = input("엑셀 파일이름을 입력해주세요")
dataframe = ytdownload(file)
os.system('pause')
